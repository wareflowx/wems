"""
Tests for enhanced Excel export functionality.

Tests cover:
- Freeze panes
- Auto-filter
- Conditional formatting for expiration dates
- Summary sheet creation
- Multi-sheet export with all enhancements
"""

import pytest
import tempfile
from pathlib import Path
from datetime import date, timedelta

from export.data_exporter import DataExporter
from employee.models import Employee, Caces, MedicalVisit, OnlineTraining


@pytest.fixture
def sample_data(db):
    """Create sample employee data in database."""
    today = date.today()

    # Create test employees
    emp1 = Employee.create(
        external_id="EMP001",
        first_name="John",
        last_name="Doe",
        email="john@example.com",
        phone="1234567890",
        current_status="Actif",
        workspace="Zone A",
        role="Cariste",
        contract_type="CDI",
        entry_date=today
    )

    emp2 = Employee.create(
        external_id="EMP002",
        first_name="Jane",
        last_name="Smith",
        email="jane@example.com",
        phone="0987654321",
        current_status="Actif",
        workspace="Zone B",
        role="Magasinier",
        contract_type="CDD",
        entry_date=today
    )

    emp3 = Employee.create(
        external_id="EMP003",
        first_name="Bob",
        last_name="Wilson",
        current_status="Inactif",
        workspace="Zone A",
        role="Préparateur",
        contract_type="Interim",
        entry_date=today
    )

    return [emp1, emp2, emp3]


@pytest.fixture
def exporter():
    """Create DataExporter instance."""
    return DataExporter()


class TestExcelExportEnhancements:
    """Test Excel export enhancements."""

    def test_export_creates_summary_sheet(self, exporter, sample_data):
        """Test that export creates a summary sheet."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            exporter.export_all_to_excel(output_path)

            import openpyxl
            wb = openpyxl.load_workbook(output_path)

            # Check that Résumé sheet exists and is first
            assert "Résumé" in wb.sheetnames
            assert wb.sheetnames[0] == "Résumé"

            # Check summary sheet has content
            ws = wb["Résumé"]
            assert ws.max_row > 1
            assert ws["A1"].value == "Tableau de Bord Employés"

            wb.close()

        finally:
            output_path.unlink(missing_ok=True)

    def test_export_creates_all_sheets(self, exporter, sample_data):
        """Test that export creates all required sheets."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            exporter.export_all_to_excel(output_path)

            import openpyxl
            wb = openpyxl.load_workbook(output_path)

            expected_sheets = ["Résumé", "Employés", "CACES", "Visites Médicales", "Formations"]
            for sheet_name in expected_sheets:
                assert sheet_name in wb.sheetnames, f"Missing sheet: {sheet_name}"

            wb.close()

        finally:
            output_path.unlink(missing_ok=True)

    def test_freeze_panes_applied(self, exporter, sample_data):
        """Test that freeze panes are applied to data sheets."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            exporter.export_all_to_excel(output_path)

            import openpyxl
            wb = openpyxl.load_workbook(output_path)

            # Check data sheets have freeze panes
            data_sheets = ["Employés", "CACES", "Visites Médicales", "Formations"]
            for sheet_name in data_sheets:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    assert ws.freeze_panes == "A2", f"Freeze panes not set in {sheet_name}"

            wb.close()

        finally:
            output_path.unlink(missing_ok=True)

    def test_auto_filter_applied(self, exporter, sample_data):
        """Test that auto-filter is applied to data sheets."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            exporter.export_all_to_excel(output_path)

            import openpyxl
            wb = openpyxl.load_workbook(output_path)

            # Check data sheets have auto-filter
            data_sheets = ["Employés", "CACES", "Visites Médicales", "Formations"]
            for sheet_name in data_sheets:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    assert ws.auto_filter.ref is not None, f"Auto-filter not set in {sheet_name}"

            wb.close()

        finally:
            output_path.unlink(missing_ok=True)

    def test_conditional_formatting_on_caces(self, exporter, sample_data):
        """Test that conditional formatting is applied to CACES sheet."""
        from datetime import date, timedelta

        # Create CACES with various expiration dates
        emp = Employee.select().first()

        # Expiring soon (< 30 days)
        Caces.create(
            employee=emp,
            kind="R489-1A",
            completion_date=date.today() - timedelta(days=300),
            expiration_date=date.today() + timedelta(days=15)
        )

        # Expiring in 30-60 days
        Caces.create(
            employee=emp,
            kind="R489-1B",
            completion_date=date.today() - timedelta(days=200),
            expiration_date=date.today() + timedelta(days=45)
        )

        # Expiring in 60-90 days
        Caces.create(
            employee=emp,
            kind="R489-3",
            completion_date=date.today() - timedelta(days=100),
            expiration_date=date.today() + timedelta(days=75)
        )

        # Expiring > 90 days
        Caces.create(
            employee=emp,
            kind="R489-4",
            completion_date=date.today() - timedelta(days=50),
            expiration_date=date.today() + timedelta(days=120)
        )

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            exporter.export_all_to_excel(output_path)

            import openpyxl
            wb = openpyxl.load_workbook(output_path)

            ws = wb["CACES"]

            # Check conditional formatting exists on column D (expiration)
            assert len(ws.conditional_formatting) > 0, "No conditional formatting applied"

            wb.close()

        finally:
            output_path.unlink(missing_ok=True)

    def test_summary_sheet_contains_statistics(self, exporter, sample_data):
        """Test that summary sheet contains employee statistics."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            exporter.export_all_to_excel(output_path)

            import openpyxl
            wb = openpyxl.load_workbook(output_path)

            ws = wb["Résumé"]

            # Check for title
            assert "Tableau de Bord Employés" in ws["A1"].value

            # Check for employee counts (row 4 is "Actif")
            assert ws["A5"].value == "Actif" or "Actif" in str(ws["A5"].value)
            assert ws["B5"].value >= 0

            wb.close()

        finally:
            output_path.unlink(missing_ok=True)

    def test_export_succeeds_without_errors(self, exporter, sample_data):
        """Test that export completes successfully."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            result = exporter.export_all_to_excel(output_path)
            assert result is True
            assert output_path.exists()

        finally:
            output_path.unlink(missing_ok=True)

    def test_export_file_size_reasonable(self, exporter, sample_data):
        """Test that exported file size is reasonable."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            exporter.export_all_to_excel(output_path)

            # File should exist and have non-zero size
            assert output_path.stat().st_size > 1000  # At least 1 KB
            assert output_path.stat().st_size < 10_000_000  # Less than 10 MB

        finally:
            output_path.unlink(missing_ok=True)

    def test_formatted_headers_exist(self, exporter, sample_data):
        """Test that headers are properly formatted."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            exporter.export_all_to_excel(output_path)

            import openpyxl
            wb = openpyxl.load_workbook(output_path)

            # Check headers in Employees sheet
            ws = wb["Employés"]

            # First row should have bold font
            for cell in ws[1]:
                assert cell.font.bold, "Header cell should be bold"

            # First row should have colored background
            header_cell = ws['A1']
            # Excel converts color codes, accept both formats
            assert header_cell.fill.start_color.rgb in ["FF366092", "00366092"]

            wb.close()

        finally:
            output_path.unlink(missing_ok=True)
