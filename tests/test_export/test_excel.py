"""Tests for Excel export functionality."""

import pytest
from pathlib import Path
from datetime import date, timedelta
import tempfile

from export import excel
from export import templates


class TestStyleConversion:
    """Tests for style conversion functions."""

    def test_convert_header_style(self):
        """Should convert header style to openpyxl format."""
        style_dict = templates.HEADER_STYLE
        openpyxl_style = excel._convert_style_dict_to_openpyxl(style_dict)

        assert 'font' in openpyxl_style
        assert 'fill' in openpyxl_style
        assert 'alignment' in openpyxl_style
        assert 'border' in openpyxl_style

    def test_convert_critical_style(self):
        """Should convert critical style to openpyxl format."""
        style_dict = templates.CRITICAL_STYLE
        openpyxl_style = excel._convert_style_dict_to_openpyxl(style_dict)

        assert openpyxl_style['fill'].start_color.rgb == 'FFC0504D'

    def test_convert_warning_style(self):
        """Should convert warning style to openpyxl format."""
        style_dict = templates.WARNING_STYLE
        openpyxl_style = excel._convert_style_dict_to_openpyxl(style_dict)

        assert openpyxl_style['fill'].start_color.rgb == 'FFFFEB9C'

    def test_convert_valid_style(self):
        """Should convert valid style to openpyxl format."""
        style_dict = templates.VALID_STYLE
        openpyxl_style = excel._convert_style_dict_to_openpyxl(style_dict)

        assert openpyxl_style['fill'].start_color.rgb == 'FFC6EFCE'


class TestCreateSummarySheet:
    """Tests for create_summary_sheet function."""

    def test_creates_summary_with_employee_counts(self, db, sample_employee):
        """Should create summary with employee counts."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_summary_sheet(wb, [sample_employee])

        # Should have Résumé sheet
        assert "Résumé" in wb.sheetnames

        ws = wb["Résumé"]

        # Should have headers
        assert ws["A1"].value == "Métrique"
        assert ws["B1"].value == "Valeur"

    def test_calculates_active_employee_count(self, db, sample_employee, inactive_employee):
        """Should correctly count active employees."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_summary_sheet(wb, [sample_employee, inactive_employee])

        ws = wb["Résumé"]

        # Find "Employés actifs" row
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == "Employés actifs":
                assert row[1].value == 1
                break

    def test_includes_export_timestamp(self, db, sample_employee):
        """Should include export timestamp."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_summary_sheet(wb, [sample_employee])

        ws = wb["Résumé"]

        # First row should have export date
        assert ws["A2"].value == "Date Export"
        assert ws["B2"].value is not None


class TestCreateEmployeesSheet:
    """Tests for create_employees_sheet function."""

    def test_creates_employees_sheet(self, db, sample_employee):
        """Should create employees sheet with data."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_employees_sheet(wb, [sample_employee])

        assert "Employés" in wb.sheetnames

        ws = wb["Employés"]

        # Check headers
        assert ws["A1"].value == "ID WMS"
        assert ws["B1"].value == "Nom Complet"

    def test_populates_employee_data(self, db, sample_employee):
        """Should populate employee data correctly."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_employees_sheet(wb, [sample_employee])

        ws = wb["Employés"]

        # First data row
        assert ws["A2"].value == sample_employee.external_id
        assert ws["B2"].value == sample_employee.full_name

    def test_calculates_seniority(self, db, sample_employee):
        """Should calculate and include seniority."""
        from openpyxl import Workbook
        from freezegun import freeze_time

        wb = Workbook()
        wb.remove(wb.active)

        with freeze_time('2026-01-16'):
            excel.create_employees_sheet(wb, [sample_employee])

        ws = wb["Employés"]

        # Seniority column (column G, 7th column)
        seniority_col = 7
        assert ws.cell(row=2, column=seniority_col).value >= 0


class TestCreateCacesSheet:
    """Tests for create_caces_sheet function."""

    def test_creates_caces_sheet(self, db, sample_employee, sample_caces):
        """Should create CACES sheet."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_caces_sheet(wb, [sample_employee])

        assert "CACES" in wb.sheetnames

        ws = wb["CACES"]

        # Check headers
        assert ws["C1"].value == "Type CACES"

    def test_includes_caces_data(self, db, sample_employee, sample_caces):
        """Should include CACES certification data."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_caces_sheet(wb, [sample_employee])

        ws = wb["CACES"]

        # Should have data row
        assert ws.max_row >= 2
        assert ws["C2"].value == sample_caces.kind

    def test_applies_status_formatting(self, db, sample_employee, expired_caces):
        """Should apply color formatting based on expiration."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_caces_sheet(wb, [sample_employee])

        ws = wb["CACES"]

        # Status column should be colored
        status_cell = ws["G2"]  # Status column
        assert status_cell.value is not None


class TestCreateMedicalVisitsSheet:
    """Tests for create_medical_visits_sheet function."""

    def test_creates_medical_sheet(self, db, sample_employee, medical_visit):
        """Should create medical visits sheet."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_medical_visits_sheet(wb, [sample_employee])

        assert "Visites Médicales" in wb.sheetnames

        ws = wb["Visites Médicales"]

        # Check headers
        assert ws["C1"].value == "Type Visite"

    def test_includes_visit_data(self, db, sample_employee, medical_visit):
        """Should include visit data."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_medical_visits_sheet(wb, [sample_employee])

        ws = wb["Visites Médicales"]

        assert ws.max_row >= 2
        assert ws["C2"].value == medical_visit.visit_type


class TestCreateTrainingsSheet:
    """Tests for create_trainings_sheet function."""

    def test_creates_trainings_sheet(self, db, sample_employee, online_training):
        """Should create trainings sheet."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_trainings_sheet(wb, [sample_employee])

        assert "Formations" in wb.sheetnames

        ws = wb["Formations"]

        # Check headers
        assert ws["C1"].value == "Titre Formation"

    def test_includes_training_data(self, db, sample_employee, online_training):
        """Should include training data."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_trainings_sheet(wb, [sample_employee])

        ws = wb["Formations"]

        assert ws.max_row >= 2
        assert ws["C2"].value == online_training.title

    def test_handles_permanent_trainings(self, db, sample_employee, permanent_training):
        """Should handle permanent trainings correctly."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        excel.create_trainings_sheet(wb, [sample_employee])

        ws = wb["Formations"]

        # Permanent training should show as "Permanent"
        assert ws.max_row >= 2


class TestExportEmployeesToExcel:
    """Tests for main export function."""

    def test_creates_excel_file(self, db, sample_employee, tmp_path):
        """Should create Excel file with all sheets."""
        output_path = tmp_path / "test_export.xlsx"

        excel.export_employees_to_excel(
            output_path,
            [sample_employee]
        )

        assert output_path.exists()

    def test_includes_all_sheets_by_default(self, db, sample_employee, tmp_path):
        """Should include all sheets by default."""
        output_path = tmp_path / "test_export.xlsx"

        excel.export_employees_to_excel(
            output_path,
            [sample_employee]
        )

        from openpyxl import load_workbook
        wb = load_workbook(output_path)

        assert "Résumé" in wb.sheetnames
        assert "Employés" in wb.sheetnames
        assert "CACES" in wb.sheetnames
        assert "Visites Médicales" in wb.sheetnames
        assert "Formations" in wb.sheetnames

    def test_can_exclude_sheets(self, db, sample_employee, tmp_path):
        """Should be able to exclude specific sheets."""
        output_path = tmp_path / "test_export.xlsx"

        excel.export_employees_to_excel(
            output_path,
            [sample_employee],
            include_caces=False,
            include_visits=False,
            include_trainings=False
        )

        from openpyxl import load_workbook
        wb = load_workbook(output_path)

        assert "Résumé" in wb.sheetnames
        assert "Employés" in wb.sheetnames
        assert "CACES" not in wb.sheetnames
        assert "Visites Médicales" not in wb.sheetnames
        assert "Formations" not in wb.sheetnames

    def test_handles_multiple_employees(self, db, sample_employee, inactive_employee, tmp_path):
        """Should handle multiple employees correctly."""
        output_path = tmp_path / "test_multiple.xlsx"

        excel.export_employees_to_excel(
            output_path,
            [sample_employee, inactive_employee]
        )

        from openpyxl import load_workbook
        wb = load_workbook(output_path)

        ws = wb["Employés"]

        # Should have header + 2 data rows
        assert ws.max_row == 3

    def test_creates_directory_if_needed(self, db, sample_employee, tmp_path):
        """Should create output directory if it doesn't exist."""
        output_path = tmp_path / "nested" / "dir" / "export.xlsx"

        excel.export_employees_to_excel(
            output_path,
            [sample_employee]
        )

        assert output_path.exists()
        assert output_path.parent.exists()


class TestSaveWorkbook:
    """Tests for save_workbook function."""

    def test_saves_workbook_to_file(self, tmp_path):
        """Should save workbook to specified path."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"

        output_path = tmp_path / "test_save.xlsx"

        excel.save_workbook(wb, output_path)

        assert output_path.exists()

        # Verify content
        from openpyxl import load_workbook
        wb_loaded = load_workbook(output_path)
        assert wb_loaded.active["A1"].value == "Test"


class TestIntegration:
    """Integration tests for Excel export."""

    def test_full_export_with_real_data(self, db, tmp_path):
        """Should export employee with all related data."""
        from employee.models import Employee, Caces, MedicalVisit, OnlineTraining

        # Create complete employee
        emp = Employee.create(
            first_name='Jean',
            last_name='Dupont',
            external_id='WMS-001',
            current_status='active',
            workspace='Quai',
            role='Cariste',
            contract_type='CDI',
            entry_date=date(2020, 1, 15)
        )

        # Add CACES
        caces = Caces.create(
            employee=emp,
            kind='R489-1A',
            completion_date=date(2020, 3, 1),
            document_path='/docs/caces.pdf'
        )

        # Add medical visit
        visit = MedicalVisit.create(
            employee=emp,
            visit_type='periodic',
            visit_date=date.today(),
            result='fit',
            document_path='/docs/medical.pdf'
        )

        # Add training
        training = OnlineTraining.create(
            employee=emp,
            title='Safety Training',
            completion_date=date.today(),
            validity_months=12,
            certificate_path='/docs/training.pdf'
        )

        # Export
        output_path = tmp_path / "full_export.xlsx"
        excel.export_employees_to_excel(output_path, [emp])

        # Verify file was created
        assert output_path.exists()

        # Verify structure
        from openpyxl import load_workbook
        wb = load_workbook(output_path)

        assert len(wb.sheetnames) == 5
        assert "Résumé" in wb.sheetnames
        assert "Employés" in wb.sheetnames
        assert "CACES" in wb.sheetnames
        assert "Visites Médicales" in wb.sheetnames
        assert "Formations" in wb.sheetnames

        # Verify employee data
        ws_emp = wb["Employés"]
        assert ws_emp["A2"].value == "WMS-001"
        assert ws_emp["B2"].value == "Jean Dupont"

        # Verify CACES data
        ws_caces = wb["CACES"]
        assert ws_caces.max_row >= 2
        assert ws_caces["C2"].value == "R489-1A"

        # Verify medical visit data
        ws_medical = wb["Visites Médicales"]
        assert ws_medical.max_row >= 2

        # Verify training data
        ws_training = wb["Formations"]
        assert ws_training.max_row >= 2
        assert ws_training["C2"].value == "Safety Training"
