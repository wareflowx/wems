"""
Tests for Data Exporter Module

Tests cover:
- JSON export for GDPR compliance
- Excel export with multiple sheets
- CSV export with proper encoding
- Edge cases (empty data, special characters, etc.)
"""

import pytest
import json
import csv
import tempfile
from pathlib import Path
from datetime import datetime, date
from unittest.mock import patch, MagicMock

import openpyxl

from src.export.data_exporter import DataExporter
from employee.models import Employee, Caces, MedicalVisit, OnlineTraining


@pytest.fixture
def temp_output_dir():
    """Create temporary directory for export files."""
    temp_dir = tempfile.mkdtemp()
    yield Path(temp_dir)

    # Cleanup
    import shutil
    shutil.rmtree(temp_dir, ignore_errors=True)


@pytest.fixture
def exporter():
    """Create DataExporter instance."""
    return DataExporter()


@pytest.fixture
def sample_employee(db):
    """Create sample employee with related data."""
    # Create employee
    employee = Employee.create(
        external_id="TEST001",
        first_name="John",
        last_name="Doe",
        email="john.doe@example.com",
        phone="1234567890",
        entry_date=date(2020, 1, 15),
        current_status="active",
        workspace="Paris",
        role="Engineer",
        contract_type="CDI",
        comment="Test employee"
    )

    # Add CACES
    caces = Caces.create(
        employee=employee,
        kind="R489-1A",
        completion_date=date(2020, 1, 1),
        expiration_date=date(2025, 1, 1),
        document_path="/docs/caces.pdf"
    )

    # Add medical visit
    visit = MedicalVisit.create(
        employee=employee,
        visit_type="initial",
        visit_date=date(2020, 1, 10),
        expiration_date=date(2025, 1, 10),
        result="fit",
        document_path="/docs/medical.pdf"
    )

    # Add training
    training = OnlineTraining.create(
        employee=employee,
        title="Safety Training",
        completion_date=date(2020, 2, 1),
        expiration_date=date(2023, 2, 1),
        document_path="/docs/training.pdf"
    )

    return employee


@pytest.fixture
def multiple_employees(db):
    """Create multiple employees for batch export."""
    employees = []

    for i in range(5):
        employee = Employee.create(
            external_id=f"TEST{i:03d}",
            first_name=f"First{i}",
            last_name=f"Last{i}",
            email=f"employee{i}@example.com",
            phone=f"12345678{i}",
            entry_date=date(2020, 1, i+1),
            current_status="active" if i % 2 == 0 else "inactive",
            workspace=f"Site{i}",
            role=f"Role{i}",
            contract_type="CDI"
        )
        employees.append(employee)

    return employees


class TestJsonExport:
    """Test JSON export functionality."""

    def test_export_employee_to_json_creates_file(self, exporter, sample_employee, temp_output_dir):
        """Test that JSON export creates a file."""
        output_path = temp_output_dir / "employee.json"

        result = exporter.export_employee_to_json(sample_employee, output_path)

        assert result is True
        assert output_path.exists()

    def test_json_export_contains_employee_data(self, exporter, sample_employee, temp_output_dir):
        """Test that JSON export contains all employee fields."""
        output_path = temp_output_dir / "employee.json"

        exporter.export_employee_to_json(sample_employee, output_path)

        with open(output_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        assert 'employee' in data
        assert data['employee']['external_id'] == sample_employee.external_id
        assert data['employee']['first_name'] == sample_employee.first_name
        assert data['employee']['last_name'] == sample_employee.last_name

    def test_json_export_contains_related_data(self, exporter, sample_employee, temp_output_dir):
        """Test that JSON export includes CACES, visits, and training."""
        output_path = temp_output_dir / "employee.json"

        exporter.export_employee_to_json(sample_employee, output_path)

        with open(output_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        assert 'caces' in data
        assert len(data['caces']) == 1
        assert data['caces'][0]['kind'] == "R489-1A"

        assert 'medical_visits' in data
        assert len(data['medical_visits']) == 1

        assert 'online_trainings' in data
        assert len(data['online_trainings']) == 1

    def test_json_export_contains_gdpr_metadata(self, exporter, sample_employee, temp_output_dir):
        """Test that JSON export includes GDPR compliance metadata."""
        output_path = temp_output_dir / "employee.json"

        exporter.export_employee_to_json(sample_employee, output_path)

        with open(output_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        assert 'export_metadata' in data
        assert 'export_date' in data['export_metadata']
        assert 'export_version' in data['export_metadata']
        assert data['export_metadata']['purpose'] == 'GDPR data portability'

    def test_json_export_handles_null_fields(self, exporter, db, temp_output_dir):
        """Test JSON export with optional fields as None."""
        employee = Employee.create(
            external_id="NULL001",
            first_name="Null",
            last_name="Test",
            email=None,
            phone=None,
            entry_date=None,
            current_status="active",
            workspace="Paris",
            role="Engineer"
        )

        output_path = temp_output_dir / "null_employee.json"

        result = exporter.export_employee_to_json(employee, output_path)

        assert result is True

        with open(output_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        assert data['employee']['email'] is None
        assert data['employee']['phone'] is None
        assert data['employee']['entry_date'] is None

    def test_json_export_uses_utf8_encoding(self, exporter, db, temp_output_dir):
        """Test that JSON export handles UTF-8 characters."""
        employee = Employee.create(
            external_id="UTF001",
            first_name="José",
            last_name="François",
            email="jose@example.com",
            current_status="active",
            workspace="Paris",
            role="Engineer"
        )

        output_path = temp_output_dir / "utf8_employee.json"

        exporter.export_employee_to_json(employee, output_path)

        # Read and verify UTF-8 encoding
        with open(output_path, 'r', encoding='utf-8') as f:
            content = f.read()

        assert "José" in content
        assert "François" in content


class TestExcelExport:
    """Test Excel export functionality."""

    def test_export_all_to_excel_creates_file(self, exporter, sample_employee, temp_output_dir):
        """Test that Excel export creates a file."""
        output_path = temp_output_dir / "employees.xlsx"

        result = exporter.export_all_to_excel(output_path)

        assert result is True
        assert output_path.exists()

    def test_excel_export_has_four_sheets(self, exporter, sample_employee, temp_output_dir):
        """Test that Excel export creates all expected sheets."""
        output_path = temp_output_dir / "employees.xlsx"

        exporter.export_all_to_excel(output_path)

        wb = openpyxl.load_workbook(output_path)
        sheet_names = wb.sheetnames

        assert "Employés" in sheet_names
        assert "CACES" in sheet_names
        assert "Visites Médicales" in sheet_names
        assert "Formations" in sheet_names
        assert len(sheet_names) == 4

        wb.close()

    def test_excel_employees_sheet_has_data(self, exporter, sample_employee, temp_output_dir):
        """Test that employees sheet contains data."""
        output_path = temp_output_dir / "employees.xlsx"

        exporter.export_all_to_excel(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Employés"]

        # Check headers
        assert ws['A1'].value == "ID Externe"
        assert ws['B1'].value == "Nom"

        # Check data row
        assert ws['A2'].value == sample_employee.external_id
        assert ws['B2'].value == sample_employee.last_name

        wb.close()

    def test_excel_employees_sheet_has_headers_styled(self, exporter, sample_employee, temp_output_dir):
        """Test that employee sheet headers are styled."""
        output_path = temp_output_dir / "employees.xlsx"

        exporter.export_all_to_excel(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Employés"]

        # Check header styling
        header_cell = ws['A1']
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb == "FF366092"

        wb.close()

    def test_excel_caces_sheet_contains_related_data(self, exporter, sample_employee, temp_output_dir):
        """Test that CACES sheet contains employee's CACES."""
        output_path = temp_output_dir / "employees.xlsx"

        exporter.export_all_to_excel(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["CACES"]

        # Check headers
        assert ws['A1'].value == "Employé"
        assert ws['B1'].value == "Kind"

        # Check data row
        assert ws['B2'].value == "R489-1A"

        wb.close()

    def test_excel_medical_sheet_contains_data(self, exporter, sample_employee, temp_output_dir):
        """Test that medical visits sheet contains data."""
        output_path = temp_output_dir / "employees.xlsx"

        exporter.export_all_to_excel(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Visites Médicales"]

        # Check data row
        assert ws['B2'].value == "Visite d'embauche"

        wb.close()

    def test_excel_training_sheet_contains_data(self, exporter, sample_employee, temp_output_dir):
        """Test that training sheet contains data."""
        output_path = temp_output_dir / "employees.xlsx"

        exporter.export_all_to_excel(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Formations"]

        # Check data row
        assert ws['B2'].value == "Safety Training"

        wb.close()

    def test_excel_export_handles_multiple_employees(self, exporter, multiple_employees, temp_output_dir):
        """Test Excel export with multiple employees."""
        output_path = temp_output_dir / "multiple_employees.xlsx"

        exporter.export_all_to_excel(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Employés"]

        # Count data rows (excluding header)
        row_count = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value is not None)

        assert row_count == 5

        wb.close()

    def test_excel_export_handles_empty_database(self, exporter, db, temp_output_dir):
        """Test Excel export with no employees."""
        output_path = temp_output_dir / "empty.xlsx"

        result = exporter.export_all_to_excel(output_path)

        assert result is True
        assert output_path.exists()


class TestCsvExport:
    """Test CSV export functionality."""

    def test_export_to_csv_creates_file(self, exporter, sample_employee, temp_output_dir):
        """Test that CSV export creates a file."""
        output_path = temp_output_dir / "employees.csv"

        result = exporter.export_to_csv(output_path)

        assert result is True
        assert output_path.exists()

    def test_csv_export_has_headers(self, exporter, sample_employee, temp_output_dir):
        """Test that CSV export has proper headers."""
        output_path = temp_output_dir / "employees.csv"

        exporter.export_to_csv(output_path)

        with open(output_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=';')
            headers = next(reader)

        assert "ID Externe" in headers
        assert "Nom" in headers
        assert "Prénom" in headers

    def test_csv_export_contains_employee_data(self, exporter, sample_employee, temp_output_dir):
        """Test that CSV export contains employee data."""
        output_path = temp_output_dir / "employees.csv"

        exporter.export_to_csv(output_path)

        with open(output_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=';')
            next(reader)  # Skip headers
            row = next(reader)

        assert row[0] == sample_employee.external_id
        assert row[1] == sample_employee.last_name
        assert row[2] == sample_employee.first_name

    def test_csv_export_uses_semicolon_delimiter(self, exporter, sample_employee, temp_output_dir):
        """Test that CSV export uses semicolon delimiter (Excel-compatible)."""
        output_path = temp_output_dir / "employees.csv"

        exporter.export_to_csv(output_path)

        with open(output_path, 'r', encoding='utf-8-sig') as f:
            content = f.read()

        # Check for semicolon delimiter
        assert ';' in content

    def test_csv_export_uses_utf8_bom_encoding(self, exporter, sample_employee, temp_output_dir):
        """Test that CSV export uses UTF-8-BOM encoding for Excel compatibility."""
        output_path = temp_output_dir / "employees.csv"

        exporter.export_to_csv(output_path)

        # Read file as binary to check BOM
        with open(output_path, 'rb') as f:
            first_bytes = f.read(3)

        # UTF-8 BOM is: EF BB BF
        assert first_bytes == b'\xef\xbb\xbf'

    def test_csv_export_handles_special_characters(self, exporter, db, temp_output_dir):
        """Test CSV export with special characters."""
        employee = Employee.create(
            external_id="SPECIAL001",
            first_name="René",
            last_name="François",
            email="rene@example.com",
            current_status="active",
            workspace="Paris",
            role="Engineer"
        )

        output_path = temp_output_dir / "special.csv"

        exporter.export_to_csv(output_path)

        with open(output_path, 'r', encoding='utf-8-sig') as f:
            content = f.read()

        assert "René" in content
        assert "François" in content

    def test_csv_export_filters_employees(self, exporter, multiple_employees, temp_output_dir):
        """Test CSV export with employee filter."""
        output_path = temp_output_dir / "filtered.csv"

        # Export only first 2 employees
        filtered_employees = multiple_employees[:2]
        exporter.export_to_csv(output_path, filtered_employees)

        with open(output_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=';')
            rows = list(reader)

        # Header + 2 data rows
        assert len(rows) == 3

    def test_csv_export_handles_null_fields(self, exporter, db, temp_output_dir):
        """Test CSV export with null/None fields."""
        employee = Employee.create(
            external_id="NULL002",
            first_name="Null",
            last_name="Test",
            email=None,
            phone=None,
            entry_date=None,
            current_status="active",
            workspace="Paris",
            role="Engineer"
        )

        output_path = temp_output_dir / "null_fields.csv"

        exporter.export_to_csv(output_path)

        with open(output_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=';')
            next(reader)  # Skip headers
            row = next(reader)

        # Null fields should be empty strings
        assert row[3] == ""  # email
        assert row[4] == ""  # phone
        assert row[5] == ""  # entry_date


class TestEdgeCases:
    """Test edge cases and error handling."""

    def test_export_employee_with_no_related_data(self, exporter, db, temp_output_dir):
        """Test exporting employee with no CACES, visits, or training."""
        employee = Employee.create(
            external_id="NOREL001",
            first_name="No",
            last_name="Relations",
            current_status="active",
            workspace="Paris",
            role="Engineer"
        )

        output_path = temp_output_dir / "no_relations.json"

        exporter.export_employee_to_json(employee, output_path)

        with open(output_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        assert len(data['caces']) == 0
        assert len(data['medical_visits']) == 0
        assert len(data['online_trainings']) == 0

    def test_json_export_handles_invalid_path(self, exporter, sample_employee):
        """Test JSON export with invalid output path."""
        invalid_path = Path("/nonexistent/directory/file.json")

        with pytest.raises(IOError):
            exporter.export_employee_to_json(sample_employee, invalid_path)

    def test_excel_export_handles_invalid_path(self, exporter, sample_employee):
        """Test Excel export with invalid output path."""
        invalid_path = Path("/nonexistent/directory/file.xlsx")

        with pytest.raises(IOError):
            exporter.export_all_to_excel(invalid_path)

    def test_csv_export_handles_invalid_path(self, exporter, sample_employee):
        """Test CSV export with invalid output path."""
        invalid_path = Path("/nonexistent/directory/file.csv")

        with pytest.raises(IOError):
            exporter.export_to_csv(invalid_path)
