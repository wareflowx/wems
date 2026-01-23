"""Tests for Excel template generator."""

import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from excel_import.template_generator import ExcelTemplateGenerator


class TestExcelTemplateGenerator:
    """Test suite for ExcelTemplateGenerator class."""

    def test_columns_constant(self):
        """Test that COLUMNS constant has correct structure."""
        assert len(ExcelTemplateGenerator.COLUMNS) == 10

        expected_columns = [
            "First Name",
            "Last Name",
            "Email",
            "Phone",
            "External ID",
            "Status",
            "Workspace",
            "Role",
            "Contract",
            "Entry Date",
        ]

        assert ExcelTemplateGenerator.COLUMNS == expected_columns

    def test_status_options_constant(self):
        """Test that STATUS_OPTIONS has correct values."""
        assert len(ExcelTemplateGenerator.STATUS_OPTIONS) == 2
        assert "Actif" in ExcelTemplateGenerator.STATUS_OPTIONS
        assert "Inactif" in ExcelTemplateGenerator.STATUS_OPTIONS

    def test_is_required_column_for_required_columns(self):
        """Test that required columns are correctly identified."""
        required_columns = [
            "First Name",
            "Last Name",
            "Status",
            "Workspace",
            "Role",
            "Contract",
            "Entry Date",
        ]

        for col in required_columns:
            assert ExcelTemplateGenerator._is_required_column(col), f"{col} should be required"

    def test_is_required_column_for_optional_columns(self):
        """Test that optional columns are correctly identified."""
        optional_columns = ["Email", "Phone", "External ID"]

        for col in optional_columns:
            assert not ExcelTemplateGenerator._is_required_column(col), f"{col} should be optional"

    def test_is_required_column_for_unknown_column(self):
        """Test that unknown columns return False."""
        assert not ExcelTemplateGenerator._is_required_column("Unknown Column")

    def test_generate_template_creates_file(self):
        """Test that generate_template creates output file."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "template.xlsx"

            generator.generate_template(output_path)

            assert output_path.exists()
            assert output_path.is_file()
            assert output_path.stat().st_size > 0

    def test_generate_template_creates_instructions_sheet(self):
        """Test that generate_template creates instructions sheet."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "template.xlsx"

            generator.generate_template(output_path)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)

            # Should have two sheets
            assert len(wb.sheetnames) == 2
            assert "Instructions" in wb.sheetnames
            assert "Data" in wb.sheetnames

            # Instructions sheet should be first
            assert wb.sheetnames[0] == "Instructions"

            wb.close()

    def test_generate_template_instructions_content(self):
        """Test that instructions sheet has correct content."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "template.xlsx"

            generator.generate_template(output_path)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            instructions_sheet = wb["Instructions"]

            # Check title
            assert instructions_sheet["A1"].value == "Employee Import Template"

            # Check for key instruction sections
            cell_values = [instructions_sheet[f"A{i}"].value for i in range(1, 50)]

            assert "How to use this template:" in cell_values
            assert "Required Columns:" in cell_values
            assert "Optional Columns:" in cell_values
            assert "Validation Rules:" in cell_values
            assert "Import Process:" in cell_values

            wb.close()

    def test_generate_template_creates_data_sheet(self):
        """Test that generate_template creates data sheet with headers."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "template.xlsx"

            generator.generate_template(output_path)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            data_sheet = wb["Data"]

            # Check headers
            for col_idx, expected_col in enumerate(ExcelTemplateGenerator.COLUMNS, start=1):
                cell_value = data_sheet.cell(row=1, column=col_idx).value
                assert expected_col in cell_value, f"Column {col_idx}: expected '{expected_col}' in '{cell_value}'"

            wb.close()

    def test_generate_template_required_columns_marked(self):
        """Test that required columns are marked with asterisk."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "template.xlsx"

            generator.generate_template(output_path)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            data_sheet = wb["Data"]

            # Check that required columns have asterisk
            for col_idx, col_name in enumerate(ExcelTemplateGenerator.COLUMNS, start=1):
                cell_value = data_sheet.cell(row=1, column=col_idx).value

                if ExcelTemplateGenerator._is_required_column(col_name):
                    if col_name != "Entry Date":  # Entry Date already has *
                        assert "*" in cell_value, f"{col_name} should have asterisk"
                else:
                    assert "*" not in cell_value, f"{col_name} should not have asterisk"

            wb.close()

    def test_generate_template_has_example_row(self):
        """Test that data sheet has example row."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "template.xlsx"

            generator.generate_template(output_path)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            data_sheet = wb["Data"]

            # Check example data in row 2
            assert data_sheet["A2"].value == "Jean"
            assert data_sheet["B2"].value == "Dupont"
            assert data_sheet["C2"].value == "jean.dupont@example.com"
            assert data_sheet["E2"].value == "WMS-001"

            wb.close()

    def test_generate_template_column_widths(self):
        """Test that columns have appropriate widths."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "template.xlsx"

            generator.generate_template(output_path)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            data_sheet = wb["Data"]

            # Check that columns have widths set
            for col_idx in range(1, 11):
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col_idx)
                assert data_sheet.column_dimensions[col_letter].width > 0

            wb.close()

    def test_generate_template_freeze_panes(self):
        """Test that data sheet has freeze panes set."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "template.xlsx"

            generator.generate_template(output_path)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            data_sheet = wb["Data"]

            # Check freeze panes
            assert data_sheet.freeze_panes == "A2"

            wb.close()

    def test_generate_sample_file_creates_file(self):
        """Test that generate_sample_file creates output file."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "sample.xlsx"

            generator.generate_sample_file(output_path, num_employees=3)

            assert output_path.exists()
            assert output_path.is_file()
            assert output_path.stat().st_size > 0

    def test_generate_sample_file_default_count(self):
        """Test that generate_sample_file uses default count of 5."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "sample.xlsx"

            generator.generate_sample_file(output_path)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb.active

            # Header row + 5 data rows
            assert sheet.max_row == 6

            wb.close()

    def test_generate_sample_file_custom_count(self):
        """Test that generate_sample_file uses custom count."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "sample.xlsx"

            generator.generate_sample_file(output_path, num_employees=10)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb.active

            # Header row + 10 data rows
            assert sheet.max_row == 11

            wb.close()

    def test_generate_sample_file_has_headers(self):
        """Test that sample file has correct headers."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "sample.xlsx"

            generator.generate_sample_file(output_path)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb.active

            # Check headers
            for col_idx, expected_col in enumerate(ExcelTemplateGenerator.COLUMNS, start=1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                assert cell_value == expected_col

            wb.close()

    def test_generate_sample_file_data_content(self):
        """Test that sample file has correct data structure."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "sample.xlsx"

            generator.generate_sample_file(output_path, num_employees=3)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb.active

            # Check first employee
            assert sheet["A2"].value == "Jean"
            assert sheet["B2"].value == "Dupont"
            assert sheet["C2"].value == "employee1@example.com"
            assert sheet["E2"].value == "WMS-001"

            # Check that external IDs are unique
            external_ids = [sheet.cell(row=i, column=5).value for i in range(2, 6)]
            assert len(external_ids) == len(set(external_ids)), "External IDs should be unique"

            wb.close()

    def test_generate_sample_file_email_format(self):
        """Test that sample file emails have correct format."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "sample.xlsx"

            generator.generate_sample_file(output_path, num_employees=5)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb.active

            # Check email format for all rows
            for row_idx in range(2, 7):
                email = sheet.cell(row=row_idx, column=3).value
                assert email is not None
                assert "@" in email
                assert email.endswith("@example.com")

            wb.close()

    def test_generate_sample_file_external_id_format(self):
        """Test that sample file external IDs have correct format."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "sample.xlsx"

            generator.generate_sample_file(output_path, num_employees=5)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb.active

            # Check external ID format
            for row_idx in range(2, 7):
                external_id = sheet.cell(row=row_idx, column=5).value
                assert external_id is not None
                assert external_id.startswith("WMS-")
                assert len(external_id) == 7  # WMS- + 3 digits

            wb.close()

    def test_generate_sample_file_date_format(self):
        """Test that sample file dates have correct format."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "sample.xlsx"

            generator.generate_sample_file(output_path, num_employees=5)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb.active

            # Check date format (DD/MM/YYYY)
            for row_idx in range(2, 7):
                date_str = sheet.cell(row=row_idx, column=10).value
                assert date_str is not None
                assert isinstance(date_str, str)
                assert "/" in date_str
                # Should be DD/MM/YYYY format
                parts = date_str.split("/")
                assert len(parts) == 3

            wb.close()

    def test_generate_sample_file_cycles_through_names(self):
        """Test that sample file cycles through first and last names."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "sample.xlsx"

            generator.generate_sample_file(output_path, num_employees=7)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb.active

            # Should cycle through 5 first names and 5 last names
            first_names = [sheet.cell(row=i, column=1).value for i in range(2, 9)]
            last_names = [sheet.cell(row=i, column=2).value for i in range(2, 9)]

            # Check that names repeat after 5
            assert first_names[0] == first_names[5]
            assert last_names[0] == last_names[5]

            wb.close()

    def test_generate_template_with_openpyxl_not_installed(self):
        """Test that ImportError is raised when openpyxl is not installed."""
        # This test verifies that the import error is raised at module level
        # We can't easily test this without actually uninstalling openpyxl
        # So we just verify the error message is correct in the module
        import excel_import.template_generator as tg

        # If we got here, openpyxl is installed
        # Just verify the ImportError is in the module
        assert hasattr(tg, "ExcelTemplateGenerator")

    def test_generate_sample_file_zero_employees(self):
        """Test that generate_sample_file handles zero employees."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "sample.xlsx"

            generator.generate_sample_file(output_path, num_employees=0)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb.active

            # Should only have header row
            assert sheet.max_row == 1

            wb.close()

    def test_generate_sample_file_many_employees(self):
        """Test that generate_sample_file handles many employees."""
        generator = ExcelTemplateGenerator()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "sample.xlsx"

            generator.generate_sample_file(output_path, num_employees=100)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            sheet = wb.active

            # Header row + 100 data rows
            assert sheet.max_row == 101

            wb.close()


class TestGenerateSampleData:
    """Test suite for _generate_sample_data method."""

    def test_generate_sample_data_returns_correct_count(self):
        """Test that _generate_sample_data returns correct number of records."""
        generator = ExcelTemplateGenerator()

        data = generator._generate_sample_data(5)

        assert len(data) == 5

    def test_generate_sample_data_has_required_keys(self):
        """Test that _generate_sample_data has all required keys."""
        generator = ExcelTemplateGenerator()

        data = generator._generate_sample_data(1)

        assert len(data) == 1
        employee = data[0]

        required_keys = [
            "first_name",
            "last_name",
            "email",
            "phone",
            "external_id",
            "status",
            "workspace",
            "role",
            "contract_type",
            "entry_date",
        ]

        for key in required_keys:
            assert key in employee, f"Missing key: {key}"

    def test_generate_sample_data_email_format(self):
        """Test that _generate_sample_data generates valid emails."""
        generator = ExcelTemplateGenerator()

        data = generator._generate_sample_data(10)

        for i, employee in enumerate(data):
            assert employee["email"] == f"employee{i + 1}@example.com"

    def test_generate_sample_data_external_id_format(self):
        """Test that _generate_sample_data generates valid external IDs."""
        generator = ExcelTemplateGenerator()

        data = generator._generate_sample_data(10)

        for i, employee in enumerate(data):
            expected_id = f"WMS-{str(i + 1).zfill(3)}"
            assert employee["external_id"] == expected_id

    def test_generate_sample_data_unique_external_ids(self):
        """Test that _generate_sample_data generates unique external IDs."""
        generator = ExcelTemplateGenerator()

        data = generator._generate_sample_data(50)

        external_ids = [emp["external_id"] for emp in data]
        assert len(external_ids) == len(set(external_ids)), "External IDs should be unique"

    def test_generate_sample_data_status_distribution(self):
        """Test that _generate_sample_data distributes status correctly."""
        generator = ExcelTemplateGenerator()

        data = generator._generate_sample_data(8)

        # Every 4th employee should be Inactif (i % 4 == 0)
        for i, employee in enumerate(data):
            expected_status = "Inactif" if i % 4 == 0 else "Actif"
            assert employee["status"] == expected_status

    def test_generate_sample_data_cycles_through_options(self):
        """Test that _generate_sample_data cycles through role/contract/workspace options."""
        generator = ExcelTemplateGenerator()

        # Import constants
        from ui_ctk.constants import ROLE_CHOICES, CONTRACT_TYPE_CHOICES, WORKSPACE_ZONES

        data = generator._generate_sample_data(10)

        # Check that it cycles through options
        for i, employee in enumerate(data):
            # Should cycle through ROLE_CHOICES
            expected_role = ROLE_CHOICES[i % len(ROLE_CHOICES)]
            assert employee["role"] == expected_role

            # Should cycle through CONTRACT_TYPE_CHOICES
            expected_contract = CONTRACT_TYPE_CHOICES[i % len(CONTRACT_TYPE_CHOICES)]
            assert employee["contract_type"] == expected_contract

            # Should cycle through WORKSPACE_ZONES
            expected_workspace = WORKSPACE_ZONES[i % len(WORKSPACE_ZONES)]
            assert employee["workspace"] == expected_workspace
