"""Create test Excel files for import testing."""

import sys
from pathlib import Path

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))


def create_test_fixtures():
    """Create all test fixtures for Excel import testing."""

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    fixtures_dir = Path(__file__).parent.parent / "tests" / "fixtures"
    fixtures_dir.mkdir(parents=True, exist_ok=True)

    print(f"Creating test fixtures in: {fixtures_dir}")
    print()

    # 1. Valid import file
    print("[1/5] Creating valid_employees.xlsx...")
    wb = Workbook()
    ws = wb.active

    headers = [
        "First Name",
        "Last Name",
        "Email",
        "Phone",
        "External ID",
        "Status",
        "Workspace",
        "Role",
        "Contract",
        "Entry Date",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    test_data = [
        [
            "Jean",
            "Dupont",
            "jean.dupont@example.com",
            "06 12 34 56 78",
            "WMS-001",
            "Actif",
            "Zone A",
            "Cariste",
            "CDI",
            "15/01/2025",
        ],
        [
            "Marie",
            "Martin",
            "marie.martin@example.com",
            "06 23 45 67 89",
            "WMS-002",
            "Actif",
            "Zone B",
            "Magasinier",
            "CDD",
            "16/01/2025",
        ],
        [
            "Pierre",
            "Bernard",
            "pierre.bernard@example.com",
            "06 34 56 78 90",
            "WMS-003",
            "Inactif",
            "Zone C",
            "PrEparateur",
            "Interim",
            "17/01/2025",
        ],
        [
            "Sophie",
            "Richard",
            "sophie.richard@example.com",
            "06 45 67 89 01",
            "WMS-004",
            "Actif",
            "Zone A",
            "Cariste",
            "Alternance",
            "18/01/2025",
        ],
        [
            "Thomas",
            "Petit",
            "thomas.petit@example.com",
            "06 56 78 90 12",
            "WMS-005",
            "Actif",
            "Zone D",
            "Magasinier",
            "CDI",
            "19/01/2025",
        ],
    ]

    for row_idx, row_data in enumerate(test_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(fixtures_dir / "valid_employees.xlsx")
    wb.close()
    print("      Created with 5 valid employee records")

    # 2. Invalid - missing required column
    print("[2/5] Creating invalid_missing_column.xlsx...")
    wb = Workbook()
    ws = wb.active

    headers = [
        "First Name",
        "Last Name",
        "Email",
        "Phone",
        "External ID",
        "Status",
        "Workspace",
        "Role",
        "Contract",
        # Entry Date is missing!
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    ws.cell(row=2, column=1, value="Jean")
    ws.cell(row=2, column=2, value="Dupont")
    ws.cell(row=2, column=6, value="Actif")
    ws.cell(row=2, column=7, value="Zone A")
    ws.cell(row=2, column=8, value="Cariste")
    ws.cell(row=2, column=9, value="CDI")

    wb.save(fixtures_dir / "invalid_missing_column.xlsx")
    wb.close()
    print("      Created - missing Entry Date column")

    # 3. Invalid - invalid data
    print("[3/5] Creating invalid_data.xlsx...")
    wb = Workbook()
    ws = wb.active

    headers = [
        "First Name",
        "Last Name",
        "Email",
        "Phone",
        "External ID",
        "Status",
        "Workspace",
        "Role",
        "Contract",
        "Entry Date",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Row 2: Invalid email
    ws.cell(row=2, column=1, value="Jean")
    ws.cell(row=2, column=2, value="Dupont")
    ws.cell(row=2, column=3, value="not-an-email")
    ws.cell(row=2, column=6, value="Actif")
    ws.cell(row=2, column=7, value="Zone A")
    ws.cell(row=2, column=8, value="Cariste")
    ws.cell(row=2, column=9, value="CDI")
    ws.cell(row=2, column=10, value="15/01/2025")

    # Row 3: Future date
    ws.cell(row=3, column=1, value="Marie")
    ws.cell(row=3, column=2, value="Martin")
    ws.cell(row=3, column=6, value="Actif")
    ws.cell(row=3, column=7, value="Zone B")
    ws.cell(row=3, column=8, value="Magasinier")
    ws.cell(row=3, column=9, value="CDD")
    ws.cell(row=3, column=10, value="01/01/2030")  # Future date

    # Row 4: Invalid workspace
    ws.cell(row=4, column=1, value="Pierre")
    ws.cell(row=4, column=2, value="Bernard")
    ws.cell(row=4, column=6, value="Actif")
    ws.cell(row=4, column=7, value="Invalid Zone")
    ws.cell(row=4, column=8, value="Cariste")
    ws.cell(row=4, column=9, value="CDI")
    ws.cell(row=4, column=10, value="15/01/2025")

    wb.save(fixtures_dir / "invalid_data.xlsx")
    wb.close()
    print("      Created with 3 rows containing various errors")

    # 4. Empty file
    print("[4/5] Creating empty_file.xlsx...")
    wb = Workbook()
    ws = wb.active
    # Only headers, no data
    headers = [
        "First Name",
        "Last Name",
        "Email",
        "Phone",
        "External ID",
        "Status",
        "Workspace",
        "Role",
        "Contract",
        "Entry Date",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    wb.save(fixtures_dir / "empty_file.xlsx")
    wb.close()
    print("      Created with headers only (no data rows)")

    # 5. Large file (100 rows)
    print("[5/5] Creating large_file.xlsx (100 rows)...")
    wb = Workbook()
    ws = wb.active

    headers = [
        "First Name",
        "Last Name",
        "Email",
        "Phone",
        "External ID",
        "Status",
        "Workspace",
        "Role",
        "Contract",
        "Entry Date",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    first_names = ["Jean", "Marie", "Pierre", "Sophie", "Thomas"]
    last_names = ["Dupont", "Martin", "Bernard", "Richard", "Petit"]
    zones = ["Zone A", "Zone B", "Zone C", "Zone D"]
    roles = ["Cariste", "Magasinier", "PrEparateur"]
    contracts = ["CDI", "CDD", "Interim", "Alternance"]

    for i in range(100):
        ws.cell(row=i + 2, column=1, value=first_names[i % len(first_names)])
        ws.cell(row=i + 2, column=2, value=last_names[i % len(last_names)])
        ws.cell(row=i + 2, column=3, value=f"employee{i + 1}@example.com")
        ws.cell(row=i + 2, column=4, value=f"06 12 34 5{i % 10:02d}")
        ws.cell(row=i + 2, column=5, value=f"WMS-{str(i + 1).zfill(3)}")
        ws.cell(row=i + 2, column=6, value="Actif" if i % 4 != 0 else "Inactif")
        ws.cell(row=i + 2, column=7, value=zones[i % len(zones)])
        ws.cell(row=i + 2, column=8, value=roles[i % len(roles)])
        ws.cell(row=i + 2, column=9, value=contracts[i % len(contracts)])
        ws.cell(row=i + 2, column=10, value=f"{15 + (i % 20):02d}/01/2025")

    wb.save(fixtures_dir / "large_file.xlsx")
    wb.close()
    print("      Created with 100 employee records")

    print()
    print("=" * 60)
    print("All test fixtures created successfully!")
    print("=" * 60)
    print()
    print("Files created:")
    print(f"  - {fixtures_dir / 'valid_employees.xlsx'}")
    print(f"  - {fixtures_dir / 'invalid_missing_column.xlsx'}")
    print(f"  - {fixtures_dir / 'invalid_data.xlsx'}")
    print(f"  - {fixtures_dir / 'empty_file.xlsx'}")
    print(f"  - {fixtures_dir / 'large_file.xlsx'}")
    print()
    print("You can now use these files for manual testing or")


if __name__ == "__main__":
    create_test_fixtures()
