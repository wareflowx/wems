"""
Data Exporter Module

Provides export functionality for employee data in multiple formats:
- JSON (GDPR-compliant data portability)
- Excel (with multiple sheets and formatting)
- CSV (simple tabular format)

Ensures GDPR compliance with:
- Article 15: Right of Access
- Article 20: Data Portability
"""

import csv
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Color
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from peewee import prefetch
from datetime import date

from employee.models import (
    Employee, Caces, MedicalVisit, OnlineTraining
)

logger = logging.getLogger(__name__)


class DataExporter:
    """Export employee data to various formats."""

    def export_employee_to_json(
        self,
        employee: Employee,
        output_path: Path
    ) -> bool:
        """
        Export single employee data to JSON (GDPR data portability).

        Args:
            employee: Employee to export
            output_path: Output file path

        Returns:
            True if successful, False otherwise

        Raises:
            IOError: If file write fails
        """
        try:
            # Load all related data using prefetch to avoid N+1 queries
            employee = Employee.get_by_id(employee.id)

            # Build complete employee data structure
            employee_data = {
                'employee': {
                    'external_id': employee.external_id,
                    'first_name': employee.first_name,
                    'last_name': employee.last_name,
                    'email': employee.email,
                    'phone': employee.phone,
                    'entry_date': employee.entry_date.isoformat() if employee.entry_date else None,
                    'current_status': employee.current_status,
                    'workspace': employee.workspace,
                    'role': employee.role,
                    'contract_type': employee.contract_type,
                },
                'caces': [
                    {
                        'kind': c.kind,
                        'completion_date': c.completion_date.isoformat() if c.completion_date else None,
                        'expiration_date': c.expiration_date.isoformat() if c.expiration_date else None,
                        'document_path': c.document_path,
                    }
                    for c in employee.caces
                ],
                'medical_visits': [
                    {
                        'visit_type': v.visit_type,
                        'visit_date': v.visit_date.isoformat() if v.visit_date else None,
                        'expiration_date': v.expiration_date.isoformat() if v.expiration_date else None,
                        'document_path': v.document_path,
                    }
                    for v in employee.medical_visits
                ],
                'online_trainings': [
                    {
                        'title': t.title,
                        'completion_date': t.completion_date.isoformat() if t.completion_date else None,
                        'expiration_date': t.expiration_date.isoformat() if t.expiration_date else None,
                        'certificate_path': t.certificate_path,
                    }
                    for t in employee.trainings
                ],
                'export_metadata': {
                    'export_date': datetime.now().isoformat(),
                    'export_version': '1.0',
                    'purpose': 'GDPR data portability',
                }
            }

            # Write to file with UTF-8 encoding
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(employee_data, f, indent=2, ensure_ascii=False)

            logger.info(f"Employee {employee.external_id} exported to JSON: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Export to JSON failed for employee {employee.external_id}: {e}")
            raise IOError(f"Failed to export employee to JSON: {e}")

    def export_all_to_excel(self, output_path: Path) -> bool:
        """
        Export all employees to Excel with multiple sheets.

        Creates four sheets:
        - Employés: Employee summary with related counts
        - CACES: Detailed CACES information
        - Visites Médicales: Medical visit details
        - Formations: Online training details

        Args:
            output_path: Output Excel file path

        Returns:
            True if successful, False otherwise

        Raises:
            IOError: If file write fails
        """
        try:
            wb = openpyxl.Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create all sheets
            self._create_employees_sheet(wb)
            self._create_caces_sheet(wb)
            self._create_medical_visits_sheet(wb)
            self._create_training_sheet(wb)
            self._create_summary_sheet(wb)

            # Save workbook
            wb.save(output_path)
            logger.info(f"All employees exported to Excel: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Export to Excel failed: {e}")
            raise IOError(f"Failed to export to Excel: {e}")

    def _create_employees_sheet(self, workbook):
        """Create employees summary sheet."""
        ws = workbook.create_sheet("Employés")

        # Headers
        headers = [
            "ID Externe", "Nom", "Prénom", "Email", "Téléphone",
            "Date Entrée", "Statut", "Zone", "Poste", "Type Contrat",
            "CACES Actifs", "Visites Actives", "Formations Actives"
        ]

        ws.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal='center')

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Load employees with related data using prefetch
        employees = list(Employee.select().order_by(Employee.last_name))

        for emp in employees:
            ws.append([
                emp.external_id,
                emp.last_name,
                emp.first_name,
                emp.email or "",
                emp.phone or "",
                emp.entry_date.isoformat() if emp.entry_date else "",
                emp.current_status,
                emp.workspace or "",
                emp.role or "",
                emp.contract_type or "",
                emp.caces.count(),
                emp.medical_visits.count(),
                emp.trainings.count(),
            ])

        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                # Skip merged cells and get column letter
                if not hasattr(cell, 'column_letter'):
                    if column_letter is None and len(column) > 0 and hasattr(column[0], 'column_letter'):
                        column_letter = column[0].column_letter
                    continue
                if column_letter is None:
                    column_letter = cell.column_letter
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Add freeze panes and auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _create_caces_sheet(self, workbook):
        """Create CACES detail sheet."""
        ws = workbook.create_sheet("CACES")

        headers = ["Employé", "Kind", "Date Complétion", "Date Expiration", "Statut"]
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Load all CACES with employee info
        caces_list = (Caces
                      .select(Caces, Employee)
                      .join(Employee)
                      .order_by(Employee.last_name))

        for c in caces_list:
            status = "Expiré" if c.is_expired else "Valide"
            ws.append([
                c.employee.full_name,
                c.kind,
                c.completion_date.isoformat() if c.completion_date else "",
                c.expiration_date.isoformat() if c.expiration_date else "",
                status,
            ])

        # Add conditional formatting for expiration dates (column D)
        self._add_expiration_conditional_formatting(ws, column_index=4)

        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                # Skip merged cells and get column letter
                if not hasattr(cell, 'column_letter'):
                    if column_letter is None and len(column) > 0 and hasattr(column[0], 'column_letter'):
                        column_letter = column[0].column_letter
                    continue
                if column_letter is None:
                    column_letter = cell.column_letter
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Add freeze panes and auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _create_medical_visits_sheet(self, workbook):
        """Create medical visits sheet."""
        ws = workbook.create_sheet("Visites Médicales")

        headers = ["Employé", "Type Visite", "Date Visite", "Date Expiration", "Statut"]
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Load all medical visits with employee info
        visits = (MedicalVisit
                  .select(MedicalVisit, Employee)
                  .join(Employee)
                  .order_by(Employee.last_name))

        for v in visits:
            status = "Expiré" if v.is_expired else "Valide"
            ws.append([
                v.employee.full_name,
                v.visit_type,
                v.visit_date.isoformat() if v.visit_date else "",
                v.expiration_date.isoformat() if v.expiration_date else "",
                status,
            ])

        # Add conditional formatting for expiration dates (column D)
        self._add_expiration_conditional_formatting(ws, column_index=4)

        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                # Skip merged cells and get column letter
                if not hasattr(cell, 'column_letter'):
                    if column_letter is None and len(column) > 0 and hasattr(column[0], 'column_letter'):
                        column_letter = column[0].column_letter
                    continue
                if column_letter is None:
                    column_letter = cell.column_letter
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Add freeze panes and auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _create_training_sheet(self, workbook):
        """Create training sheet."""
        ws = workbook.create_sheet("Formations")

        headers = ["Employé", "Title", "Date Complétion", "Date Expiration", "Statut"]
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Load all trainings with employee info
        trainings = (OnlineTraining
                     .select(OnlineTraining, Employee)
                     .join(Employee)
                     .order_by(Employee.last_name))

        for t in trainings:
            status = "Expiré" if t.is_expired else "Valide"
            ws.append([
                t.employee.full_name,
                t.title,
                t.completion_date.isoformat() if t.completion_date else "",
                t.expiration_date.isoformat() if t.expiration_date else "",
                status,
            ])

        # Add conditional formatting for expiration dates (column D)
        self._add_expiration_conditional_formatting(ws, column_index=4)

        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                # Skip merged cells and get column letter
                if not hasattr(cell, 'column_letter'):
                    if column_letter is None and len(column) > 0 and hasattr(column[0], 'column_letter'):
                        column_letter = column[0].column_letter
                    continue
                if column_letter is None:
                    column_letter = cell.column_letter
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Add freeze panes and auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _add_expiration_conditional_formatting(self, worksheet, column_index: int):
        """
        Add conditional formatting to expiration date column.

        Color coding:
        - Red (< 30 days): Critical
        - Orange (30-60 days): Warning
        - Yellow (60-90 days): Info
        - Green (> 90 days or no date): OK

        Args:
            worksheet: Excel worksheet
            column_index: 1-based column index for expiration dates
        """
        if worksheet.max_row < 2:
            return

        column_letter = get_column_letter(column_index)
        range_string = f"{column_letter}2:{column_letter}{worksheet.max_row}"

        from openpyxl.formatting.rule import CellIsRule

        # Red for expired or < 30 days
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_rule = CellIsRule(
            operator='lessThan',
            formula=['=TODAY()+30'],
            fill=red_fill
        )
        worksheet.conditional_formatting.add(range_string, red_rule)

        # Orange for 30-60 days
        orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        orange_rule = CellIsRule(
            operator='between',
            formula=['=TODAY()+30', '=TODAY()+60'],
            fill=orange_fill
        )
        worksheet.conditional_formatting.add(range_string, orange_rule)

        # Yellow for 60-90 days
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        yellow_rule = CellIsRule(
            operator='between',
            formula=['=TODAY()+60', '=TODAY()+90'],
            fill=yellow_fill
        )
        worksheet.conditional_formatting.add(range_string, yellow_rule)

        # Green for > 90 days
        green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        green_rule = CellIsRule(
            operator='greaterThan',
            formula=['=TODAY()+90'],
            fill=green_fill
        )
        worksheet.conditional_formatting.add(range_string, green_rule)

    def _create_summary_sheet(self, workbook):
        """Create summary statistics sheet."""
        ws = workbook.create_sheet("Résumé", 0)  # First sheet

        # Title
        ws.append(["Tableau de Bord Employés"])
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        ws.append([])  # Empty row

        # Employee counts by status
        ws.append(["Employés par Statut"])
        ws.append(["Statut", "Nombre"])

        active_count = Employee.select().where(Employee.current_status == "Actif").count()
        inactive_count = Employee.select().where(Employee.current_status == "Inactif").count()
        total_count = Employee.select().count()

        ws.append(["Actif", active_count])
        ws.append(["Inactif", inactive_count])
        ws.append(["Total", total_count])

        ws.append([])  # Empty row

        # Employee counts by workspace
        ws.append(["Employés par Zone"])
        ws.append(["Zone", "Nombre"])

        for workspace in Employee.select(Employee.workspace).distinct():
            if workspace.workspace:
                count = Employee.select().where(Employee.workspace == workspace.workspace).count()
                ws.append([workspace.workspace, count])

        ws.append([])  # Empty row

        # Employee counts by role
        ws.append(["Employés par Poste"])
        ws.append(["Poste", "Nombre"])

        for role in Employee.select(Employee.role).distinct():
            if role.role:
                count = Employee.select().where(Employee.role == role.role).count()
                ws.append([role.role, count])

        ws.append([])  # Empty row

        # Certification counts
        ws.append(["Certifications"])
        ws.append(["Type", "Total", "Expirés", "Valides"])

        # CACES
        total_caces = Caces.select().count()
        expired_caces = Caces.select().where(Caces.is_expired == True).count()
        valid_caces = total_caces - expired_caces
        ws.append(["CACES", total_caces, expired_caces, valid_caces])

        # Medical visits
        total_visits = MedicalVisit.select().count()
        expired_visits = MedicalVisit.select().where(MedicalVisit.is_expired == True).count()
        valid_visits = total_visits - expired_visits
        ws.append(["Visites Médicales", total_visits, expired_visits, valid_visits])

        # Training
        total_training = OnlineTraining.select().count()
        expired_training = OnlineTraining.select().where(OnlineTraining.is_expired == True).count()
        valid_training = total_training - expired_training
        ws.append(["Formations", total_training, expired_training, valid_training])

        ws.append([])  # Empty row
        ws.append([f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"])

        # Apply formatting
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.row in [1, 3, 8, 12, 17, 21]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                # Skip merged cells and get column letter
                if not hasattr(cell, 'column_letter'):
                    if column_letter is None and len(column) > 0 and hasattr(column[0], 'column_letter'):
                        column_letter = column[0].column_letter
                    continue
                if column_letter is None:
                    column_letter = cell.column_letter
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def export_to_csv(
        self,
        output_path: Path,
        employees: Optional[List[Employee]] = None
    ) -> bool:
        """
        Export employee data to CSV.

        Args:
            output_path: Output CSV file path
            employees: List of employees to export (all if None)

        Returns:
            True if successful, False otherwise

        Raises:
            IOError: If file write fails
        """
        try:
            if employees is None:
                employees = list(Employee.select().order_by(Employee.last_name))

            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')

                # Headers
                headers = [
                    "ID Externe", "Nom", "Prénom", "Email", "Téléphone",
                    "Date Entrée", "Statut", "Zone", "Poste", "Type Contrat"
                ]
                writer.writerow(headers)

                # Data rows
                for emp in employees:
                    writer.writerow([
                        emp.external_id,
                        emp.last_name,
                        emp.first_name,
                        emp.email or "",
                        emp.phone or "",
                        emp.entry_date.isoformat() if emp.entry_date else "",
                        emp.current_status,
                        emp.workspace or "",
                        emp.role or "",
                        emp.contract_type or "",
                    ])

            logger.info(f"{len(employees)} employees exported to CSV: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Export to CSV failed: {e}")
            raise IOError(f"Failed to export to CSV: {e}")
