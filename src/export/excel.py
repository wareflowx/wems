"""Excel generation with formatting."""

from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")

from employee import calculations
from employee.models import Employee
from export import templates

# ========== STYLE CONVERSION ==========


def _convert_style_dict_to_openpyxl(style_dict: dict[str, Any]) -> dict:
    """
    Convert our style dict format to openpyxl style objects.

    Args:
        style_dict: Our style dictionary format

    Returns:
        Dictionary with openpyxl style objects
    """
    openpyxl_style = {}

    # Convert font
    if "font" in style_dict:
        font_props = style_dict["font"]
        openpyxl_style["font"] = Font(
            bold=font_props.get("bold", False), size=font_props.get("size", 11), color=font_props.get("color", "000000")
        )

    # Convert fill
    if "fill" in style_dict:
        fill_props = style_dict["fill"]
        # Add 'FF' prefix if not present for openpyxl (expects 8 chars with alpha)
        fg_color = fill_props.get("fgColor", "FFFFFF")
        if len(fg_color) == 6:  # RGB without alpha
            fg_color = "FF" + fg_color
        # If already 8 chars, use as-is

        openpyxl_style["fill"] = PatternFill(
            start_color=fg_color, end_color=fg_color, fill_type=fill_props.get("fill_type", "solid")
        )

    # Convert alignment
    if "alignment" in style_dict:
        align_props = style_dict["alignment"]
        openpyxl_style["alignment"] = Alignment(
            horizontal=align_props.get("horizontal", "left"), vertical=align_props.get("vertical", "center")
        )

    # Convert border
    if "border" in style_dict:
        border_props = style_dict["border"]
        thin_side = Side(border_style="thin")

        openpyxl_style["border"] = Border(
            top=thin_side if "top" in border_props else None,
            left=thin_side if "left" in border_props else None,
            bottom=thin_side if "bottom" in border_props else None,
            right=thin_side if "right" in border_props else None,
        )

    return openpyxl_style


def _apply_style_to_cell(cell, style_dict: dict[str, Any]) -> None:
    """
    Apply style dictionary to a cell.

    Args:
        cell: openpyxl cell object
        style_dict: Our style dictionary format
    """
    openpyxl_style = _convert_style_dict_to_openpyxl(style_dict)

    if "font" in openpyxl_style:
        cell.font = openpyxl_style["font"]
    if "fill" in openpyxl_style:
        cell.fill = openpyxl_style["fill"]
    if "alignment" in openpyxl_style:
        cell.alignment = openpyxl_style["alignment"]
    if "border" in openpyxl_style:
        cell.border = openpyxl_style["border"]


# ========== MAIN EXPORT FUNCTION ==========


def export_employees_to_excel(
    output_path: Path,
    employees: list[Employee],
    include_caces: bool = True,
    include_visits: bool = True,
    include_trainings: bool = True,
) -> None:
    """
    Export employees to Excel file with multiple sheets.

    Creates workbook with sheets:
    - Résumé: Overview statistics
    - Employés: Employee data
    - CACES: Certification details (if include_caces)
    - Visites Médicales: Visit details (if include_visits)
    - Formations: Training details (if include_trainings)

    Args:
        output_path: Where to save the Excel file
        employees: List of Employee objects to export
        include_caces: Include CACES sheet
        include_visits: Include medical visits sheet
        include_trainings: Include training sheet

    Raises:
        IOError: If file cannot be written
        PermissionError: If file is locked by another process

    Example:
        >>> employees = Employee.select()
        >>> export_employees_to_excel(
        ...     Path("export.xlsx"),
        ...     list(employees)
        ... )
    """
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheets
    create_summary_sheet(wb, employees)
    create_employees_sheet(wb, employees)

    if include_caces:
        create_caces_sheet(wb, employees)

    if include_visits:
        create_medical_visits_sheet(wb, employees)

    if include_trainings:
        create_trainings_sheet(wb, employees)

    # Save workbook
    save_workbook(wb, output_path)


# ========== SHEET CREATION FUNCTIONS ==========


def create_summary_sheet(workbook: Workbook, employees: list[Employee]) -> None:
    """
    Create summary sheet with key metrics.

    Metrics:
    - Total employees
    - Active employees
    - Expiring items (by category)
    - Unfit employees

    Args:
        workbook: Workbook object to add sheet to
        employees: List of employees
    """
    ws = workbook.create_sheet("Résumé", 0)

    # Calculate statistics
    total_employees = len(employees)
    active_employees = len([e for e in employees if e.current_status == "active"])

    # Count expiring items
    expired_caces = 0
    critical_caces = 0
    expired_visits = 0
    critical_visits = 0
    expired_trainings = 0
    critical_trainings = 0

    for emp in employees:
        # CACES
        for caces in emp.caces:
            if caces.is_expired:
                expired_caces += 1
            elif caces.status == "critical":
                critical_caces += 1

        # Medical visits
        for visit in emp.medical_visits:
            if visit.is_expired:
                expired_visits += 1
            elif visit.days_until_expiration < 30:
                critical_visits += 1

        # Trainings
        for training in emp.trainings:
            if training.expires:
                if training.is_expired:
                    expired_trainings += 1
                elif training.days_until_expiration < 30:
                    critical_trainings += 1

    unfit_count = len([v for e in employees for v in e.medical_visits if v.result == "unfit"])

    # Define metrics
    metrics_data = [
        ("Date Export", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Employés", ""),
        ("Total employés", total_employees),
        ("Employés actifs", active_employees),
        ("Employés inactifs", total_employees - active_employees),
        ("", ""),
        ("CACES", ""),
        ("CACES expirés", expired_caces),
        ("CACES critiques (< 30 j)", critical_caces),
        ("", ""),
        ("Visites Médicales", ""),
        ("Visites expirées", expired_visits),
        ("Visites critiques (< 30 j)", critical_visits),
        ("Employés inaptes", unfit_count),
        ("", ""),
        ("Formations", ""),
        ("Formations expirées", expired_trainings),
        ("Formations critiques (< 30 j)", critical_trainings),
    ]

    # Write headers
    headers = templates.get_headers_for_columns(templates.SUMMARY_COLUMNS)
    ws.append(headers)
    _apply_style_to_row(ws, 1, templates.HEADER_STYLE)

    # Write data
    for row_idx, (metric, value) in enumerate(metrics_data, start=2):
        ws.cell(row=row_idx, column=1, value=metric)
        ws.cell(row=row_idx, column=2, value=value)

        # Apply styles
        _apply_style_to_cell(ws.cell(row=row_idx, column=1), templates.METRIC_STYLE)
        _apply_style_to_cell(ws.cell(row=row_idx, column=2), templates.VALUE_STYLE)

    # Set column widths
    ws.column_dimensions["A"].width = templates.SUMMARY_COLUMNS[0]["width"]
    ws.column_dimensions["B"].width = templates.SUMMARY_COLUMNS[1]["width"]


def create_employees_sheet(workbook: Workbook, employees: list[Employee]) -> None:
    """
    Create employee data sheet with headers.

    Args:
        workbook: Workbook object to add sheet to
        employees: List of employees
    """
    ws = workbook.create_sheet("Employés")

    # Write headers
    headers = templates.get_headers_for_columns(templates.EMPLOYEE_COLUMNS)
    ws.append(headers)
    _apply_style_to_row(ws, 1, templates.HEADER_STYLE)

    # Write employee data
    keys = templates.get_keys_for_columns(templates.EMPLOYEE_COLUMNS)

    for emp in employees:
        row_data = []
        for key in keys:
            if key == "full_name":
                value = emp.full_name
            elif key == "seniority":
                value = calculations.calculate_seniority(emp)
            elif key == "status":
                score = calculations.calculate_compliance_score(emp)
                status = calculations.get_compliance_status(emp)
                # Show status text
                status_map = {"critical": "Critique", "warning": "Attention", "compliant": "Conforme"}
                value = status_map.get(status, "Inconnu")
            else:
                value = getattr(emp, key, "")

            row_data.append(value)

        ws.append(row_data)

        # Apply conditional formatting to status column
        status_col_idx = len(keys)
        status_cell = ws.cell(row=ws.max_row, column=status_col_idx)
        if "status" in keys:
            status_value = row_data[-1]
            style = templates.get_style_for_status(status_value)
            _apply_style_to_cell(status_cell, style)

    # Set column widths
    for idx, width in enumerate(templates.get_column_widths(templates.EMPLOYEE_COLUMNS), 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"


def create_caces_sheet(workbook: Workbook, employees: list[Employee]) -> None:
    """
    Create CACES details sheet.

    One row per CACES, grouped by employee.
    Apply conditional formatting (red for expired, orange for critical).

    Args:
        workbook: Workbook object
        employees: List of employees
    """
    ws = workbook.create_sheet("CACES")

    # Write headers
    headers = templates.get_headers_for_columns(templates.CACES_COLUMNS)
    ws.append(headers)
    _apply_style_to_row(ws, 1, templates.HEADER_STYLE)

    # Write CACES data
    for emp in employees:
        for caces in emp.caces:
            # Determine status
            if caces.is_expired:
                status = "expired"
            elif caces.status == "critical":
                status = "critical"
            else:
                status = "valid"

            row_data = [
                emp.external_id,
                emp.full_name,
                caces.kind,
                caces.completion_date,
                caces.expiration_date,
                caces.days_until_expiration,
                status.capitalize(),
                str(caces.document_path) if caces.document_path else "",
            ]

            ws.append(row_data)

            # Apply status style
            status_col_idx = 7
            status_cell = ws.cell(row=ws.max_row, column=status_col_idx)
            style = templates.get_style_for_status(status)
            _apply_style_to_cell(status_cell, style)

    # Set column widths
    for idx, width in enumerate(templates.get_column_widths(templates.CACES_COLUMNS), 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"


def create_medical_visits_sheet(workbook: Workbook, employees: list[Employee]) -> None:
    """
    Create medical visits sheet.

    Apply conditional formatting based on fitness status.

    Args:
        workbook: Workbook object
        employees: List of employees
    """
    ws = workbook.create_sheet("Visites Médicales")

    # Write headers
    headers = templates.get_headers_for_columns(templates.MEDICAL_COLUMNS)
    ws.append(headers)
    _apply_style_to_row(ws, 1, templates.HEADER_STYLE)

    # Write medical visit data
    for emp in employees:
        for visit in emp.medical_visits:
            # Determine status
            if visit.is_expired:
                status = "expired"
            elif visit.days_until_expiration < 30:
                status = "critical"
            else:
                status = "valid"

            row_data = [
                emp.external_id,
                emp.full_name,
                visit.visit_type,
                visit.visit_date,
                visit.expiration_date,
                visit.days_until_expiration,
                visit.result,
                status.capitalize(),
                str(visit.document_path) if visit.document_path else "",
            ]

            ws.append(row_data)

            # Apply styles to result and status columns
            result_col_idx = 7
            status_col_idx = 8

            result_cell = ws.cell(row=ws.max_row, column=result_col_idx)
            status_cell = ws.cell(row=ws.max_row, column=status_col_idx)

            # Style based on result
            result_style = templates.get_style_for_status(visit.result)
            _apply_style_to_cell(result_cell, result_style)

            # Style based on expiration status
            status_style = templates.get_style_for_status(status)
            _apply_style_to_cell(status_cell, status_style)

    # Set column widths
    for idx, width in enumerate(templates.get_column_widths(templates.MEDICAL_COLUMNS), 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"


def create_trainings_sheet(workbook: Workbook, employees: list[Employee]) -> None:
    """
    Create trainings sheet.

    Apply conditional formatting for expiring trainings.

    Args:
        workbook: Workbook object
        employees: List of employees
    """
    ws = workbook.create_sheet("Formations")

    # Write headers
    headers = templates.get_headers_for_columns(templates.TRAINING_COLUMNS)
    ws.append(headers)
    _apply_style_to_row(ws, 1, templates.HEADER_STYLE)

    # Write training data
    for emp in employees:
        for training in emp.trainings:
            # Skip permanent trainings for status
            if not training.expires:
                status = "valid"
                days_until = None
            else:
                if training.is_expired:
                    status = "expired"
                elif training.days_until_expiration < 30:
                    status = "critical"
                else:
                    status = "valid"
                days_until = training.days_until_expiration

            row_data = [
                emp.external_id,
                emp.full_name,
                training.title,
                training.completion_date,
                training.expiration_date if training.expires else "Permanent",
                days_until,
                status.capitalize() if training.expires else "Permanent",
                str(training.certificate_path) if training.certificate_path else "",
            ]

            ws.append(row_data)

            # Apply status style (skip for permanent trainings)
            if training.expires:
                status_col_idx = 7
                status_cell = ws.cell(row=ws.max_row, column=status_col_idx)
                style = templates.get_style_for_status(status)
                _apply_style_to_cell(status_cell, style)

    # Set column widths
    for idx, width in enumerate(templates.get_column_widths(templates.TRAINING_COLUMNS), 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"


# ========== HELPER FUNCTIONS ==========


def save_workbook(workbook: Workbook, path: Path) -> None:
    """
    Save workbook to file with error handling.

    Args:
        workbook: Workbook to save
        path: Where to save the file

    Raises:
        IOError: If file cannot be written
        PermissionError: If file is locked
    """
    # Ensure parent directory exists
    path.parent.mkdir(parents=True, exist_ok=True)

    # Try to save
    try:
        workbook.save(path)
    except PermissionError:
        # Try with timestamp suffix
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_path = path.parent / f"{path.stem}_{timestamp}{path.suffix}"
        workbook.save(new_path)
        raise PermissionError(
            f"Could not write to {path}. Saved as {new_path} instead. Close the file if it's open in Excel."
        )


def _apply_style_to_row(worksheet, row_idx: int, style_dict: dict[str, Any]) -> None:
    """
    Apply style to all cells in a row.

    Args:
        worksheet: Worksheet object
        row_idx: Row number (1-based)
        style_dict: Style dictionary to apply
    """
    for cell in worksheet[row_idx]:
        _apply_style_to_cell(cell, style_dict)
