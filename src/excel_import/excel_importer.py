"""Excel import logic for bulk employee import."""

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl is required for Excel import. Install it with: pip install openpyxl")

from database.connection import database
from employee.models import Employee
from utils.validation import InputValidator, ValidationError
from ui_ctk.constants import (
    CONTRACT_TYPE_CHOICES,
    ROLE_CHOICES,
    WORKSPACE_ZONES,
)


@dataclass
class ImportError:
    """
    Import error for a specific row/cell.

    Attributes:
        row_num: Row number in Excel file (1-indexed)
        column: Column name where error occurred
        value: The invalid value
        error_type: Type of error (required, format, duplicate, validation)
        message: Human-readable error message
        severity: Error severity (critical, warning, info)
    """

    row_num: int
    column: str
    value: Any
    error_type: str
    message: str
    severity: str = "warning"

    def __str__(self) -> str:
        return f"Row {self.row_num}, {self.column}: {self.message}"


@dataclass
class ImportResult:
    """
    Result of Excel import operation.

    Attributes:
        total_rows: Total number of rows processed
        successful: Number of successfully imported employees
        failed: Number of failed rows
        skipped: Number of skipped rows (duplicates, etc.)
        errors: List of ImportError objects
        duration: Import duration in seconds
        file_path: Path to imported file
    """

    total_rows: int = 0
    successful: int = 0
    failed: int = 0
    skipped: int = 0
    errors: List[ImportError] = field(default_factory=list)
    duration: float = 0.0
    file_path: Optional[Path] = None

    @property
    def has_errors(self) -> bool:
        """Check if import had any errors."""
        return len(self.errors) > 0

    @property
    def success_rate(self) -> float:
        """Calculate success rate as percentage."""
        if self.total_rows == 0:
            return 0.0
        return (self.successful / self.total_rows) * 100


class ExcelImporter:
    """
    Handles bulk import of employees from Excel files.

    Features:
    - Excel file parsing with openpyxl
    - Column validation and mapping
    - Data validation per row
    - Batch processing with transactions
    - Duplicate detection
    - Detailed error reporting
    - Progress tracking support
    """

    # Required columns in Excel file
    REQUIRED_COLUMNS = ["First Name", "Last Name", "Status", "Workspace", "Role", "Contract", "Entry Date"]

    # Optional columns
    OPTIONAL_COLUMNS = ["Email", "Phone", "External ID"]

    # Batch size for transaction commits
    BATCH_SIZE = 100

    def __init__(self, file_path: Path):
        """
        Initialize importer with Excel file path.

        Args:
            file_path: Path to Excel file (.xlsx)
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.headers = []
        self.data_rows = []

    def validate_file(self) -> Tuple[bool, Optional[str]]:
        """
        Validate Excel file before processing.

        Checks:
        - File exists
        - .xlsx extension
        - Can be opened by openpyxl
        - Has at least one worksheet
        - Contains required columns

        Returns:
            Tuple of (is_valid, error_message)
        """
        # Check file exists
        if not self.file_path.exists():
            return False, f"File not found: {self.file_path}"

        # Check extension
        if self.file_path.suffix.lower() != ".xlsx":
            return False, "File must be .xlsx format"

        # Try to open file
        try:
            self.workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        except Exception as e:
            return False, f"Cannot open file: {str(e)}"

        # Check worksheets
        if len(self.workbook.sheetnames) == 0:
            return False, "File contains no worksheets"

        # Use first worksheet
        self.worksheet = self.workbook.active

        # Check if worksheet has data
        if self.worksheet.max_row < 1:
            return False, "File is empty (no data rows)"

        # Load headers from first row
        self.headers = [
            self._get_cell_value(row_idx=1, col_idx=col_idx) for col_idx in range(1, self.worksheet.max_column + 1)
        ]

        # Check for required columns
        missing_columns = set(self.REQUIRED_COLUMNS) - set(self.headers)
        if missing_columns:
            return False, f"Missing required columns: {', '.join(missing_columns)}"

        return True, None

    def _get_cell_value(self, row_idx: int, col_idx: int) -> Optional[str]:
        """
        Get cell value as string.

        Args:
            row_idx: Row index (1-indexed)
            col_idx: Column index (1-indexed)

        Returns:
            Cell value as string or None if empty
        """
        cell = self.worksheet.cell(row_idx, col_idx)
        if cell.value is None:
            return None

        value = str(cell.value).strip()
        return value if value else None

    def close(self):
        """
        Close the workbook and release resources.

        Should be called when done with the importer to prevent file locks.
        """
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def parse_file(self) -> List[Dict[str, Any]]:
        """
        Parse Excel file and return list of row dictionaries.

        Returns:
            List of dictionaries, each containing:
            {
                'row_num': int,
                'data': {column_name: cell_value, ...},
                'raw_row': dict of raw values
            }
        """
        if not self.worksheet:
            raise RuntimeError("File not validated. Call validate_file() first.")

        rows = []

        # Start from row 2 (skip header)
        for row_idx in range(2, self.worksheet.max_row + 1):
            row_data = {}
            raw_row = {}

            for col_idx, header in enumerate(self.headers, start=1):
                value = self._get_cell_value(row_idx, col_idx)
                if value:
                    row_data[header] = value
                raw_row[header] = value

            # Only include rows that have at least some data
            if row_data:
                rows.append({"row_num": row_idx, "data": row_data, "raw_row": raw_row})

        self.data_rows = rows
        return rows

    def preview(self, max_rows: int = 3) -> Dict[str, Any]:
        """
        Generate preview of Excel data.

        Args:
            max_rows: Maximum number of sample rows to show

        Returns:
            Dictionary with preview data:
            {
                'total_rows': int,
                'columns': List[str],
                'sample_data': List[Dict],
                'detected_issues': List[str]
            }
        """
        if not self.data_rows:
            self.parse_file()

        # Get sample rows
        sample_data = self.data_rows[:max_rows]

        # Detect issues
        issues = []

        # Check for empty required fields in sample
        for row in sample_data:
            for col in self.REQUIRED_COLUMNS:
                if col not in row["data"] or not row["data"][col]:
                    issues.append(f"Row {row['row_num']}: Missing '{col}'")

        return {
            "total_rows": len(self.data_rows),
            "columns": self.headers,
            "sample_data": sample_data,
            "detected_issues": issues,
        }

    def import_employees(self, progress_callback: Optional[Callable[[int, int], None]] = None) -> ImportResult:
        """
        Import employees from parsed Excel data.

        Processes all parsed rows in batches with transaction support.
        Detects duplicates and collects errors.

        Args:
            progress_callback: Optional callback for progress updates
                              Called with (current_row, total_rows)

        Returns:
            ImportResult with detailed statistics
        """
        if not self.data_rows:
            self.parse_file()

        start_time = datetime.now()
        result = ImportResult(file_path=self.file_path)
        result.total_rows = len(self.data_rows)

        if progress_callback:
            progress_callback(0, result.total_rows)

        # Process in batches
        for batch_start in range(0, result.total_rows, self.BATCH_SIZE):
            batch_end = min(batch_start + self.BATCH_SIZE, result.total_rows)
            batch = self.data_rows[batch_start:batch_end]

            # Process batch with transaction
            try:
                with database.atomic():
                    for row_info in batch:
                        success, error = self._import_single_row(row_info)

                        if success:
                            result.successful += 1
                        else:
                            result.failed += 1
                            if error:
                                result.errors.append(error)

                # Update progress
                if progress_callback:
                    progress_callback(batch_end, result.total_rows)

            except Exception as e:
                # Transaction was rolled back
                # Count all rows in batch as failed
                result.failed += len(batch)
                for row_info in batch:
                    result.errors.append(
                        ImportError(
                            row_num=row_info["row_num"],
                            column="multiple",
                            value="N/A",
                            error_type="database",
                            message=f"Batch failed: {str(e)}",
                            severity="critical",
                        )
                    )
                break

        # Calculate duration
        result.duration = (datetime.now() - start_time).total_seconds()

        # Close workbook
        if self.workbook:
            self.workbook.close()

        return result

    def _import_single_row(self, row_info: Dict[str, Any]) -> Tuple[bool, Optional[ImportError]]:
        """
        Import a single row of employee data.

        Args:
            row_info: Row dictionary from parse_file()

        Returns:
            Tuple of (success, error_object)
        """
        row_num = row_info["row_num"]
        data = row_info["data"]

        try:
            # Map row data to employee fields
            employee_data = self._map_row_to_employee(data)

            # Validate data
            error = self._validate_row(row_num, employee_data)
            if error:
                return False, error

            # Check for duplicate external_id
            if employee_data.get("external_id"):
                dup_error = self._check_duplicate_external_id(employee_data["external_id"])
                if dup_error:
                    return False, dup_error

            # Create employee
            employee = Employee.create(**employee_data)
            print(f"[OK] Imported: {employee.full_name} (row {row_num})")
            return True, None

        except Exception as e:
            return False, ImportError(
                row_num=row_num,
                column="general",
                value=str(data),
                error_type="exception",
                message=str(e),
                severity="critical",
            )

    def _map_row_to_employee(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Map Excel row data to Employee model fields.

        Args:
            row_data: Dictionary from Excel row

        Returns:
            Dictionary with Employee model fields
        """
        # Extract and transform fields
        employee_data = {}

        # Name fields
        employee_data["first_name"] = self._clean_string(row_data.get("First Name"))
        employee_data["last_name"] = self._clean_string(row_data.get("Last Name"))

        # Contact fields (optional)
        employee_data["email"] = self._clean_string(row_data.get("Email"))
        employee_data["phone"] = self._clean_string(row_data.get("Phone"))

        # External ID (optional)
        employee_data["external_id"] = self._clean_string(row_data.get("External ID"))

        # Status - Map French to English
        status_str = self._clean_string(row_data.get("Status", "Actif"))
        employee_data["current_status"] = "active" if status_str == "Actif" else "inactive"

        # Workspace
        employee_data["workspace"] = self._clean_string(row_data.get("Workspace"))

        # Role
        employee_data["role"] = self._clean_string(row_data.get("Role"))

        # Contract type
        employee_data["contract_type"] = self._clean_string(row_data.get("Contract"))

        # Entry date
        employee_data["entry_date"] = self._parse_date(row_data.get("Entry Date"))

        return employee_data

    def _validate_row(self, row_num: int, employee_data: Dict[str, Any]) -> Optional[ImportError]:
        """
        Validate a single row of employee data using InputValidator.

        Args:
            row_num: Row number (for error reporting)
            employee_data: Mapped employee data

        Returns:
            ImportError if invalid, None if valid
        """
        try:
            # Use comprehensive validation framework
            validated_data = InputValidator.validate_employee_data(employee_data)
            return None

        except ValidationError as e:
            # Map ValidationError to ImportError
            return ImportError(
                row_num=row_num,
                column=e.field,
                value=e.value,
                error_type="validation",
                message=e.message,
            )

    def _check_duplicate_external_id(self, external_id: str) -> Optional[ImportError]:
        """
        Check if external_id already exists in database.

        Args:
            external_id: External ID to check

        Returns:
            ImportError if duplicate, None otherwise
        """
        try:
            existing = Employee.select().where(Employee.external_id == external_id).first()
            if existing:
                return ImportError(
                    row_num=0,  # Will be set by caller
                    column="External ID",
                    value=external_id,
                    error_type="duplicate",
                    message=f"External ID '{external_id}' already exists (employee: {existing.full_name})",
                )
        except Exception:
            pass

        return None

    @staticmethod
    def _clean_string(value: Optional[str]) -> Optional[str]:
        """
        Clean string value.

        - Trim whitespace
        - Capitalize first letter
        - Return None if empty

        Args:
            value: String value to clean

        Returns:
            Cleaned string or None
        """
        if not value:
            return None

        value = value.strip()
        if not value:
            return None

        # Capitalize first letter
        return value[0].upper() + value[1:].lower() if len(value) > 1 else value.upper()

    @staticmethod
    def _parse_date(date_str: Optional[str]) -> Optional[date]:
        """
        Parse date from multiple formats.

        Tries in order:
        1. DD/MM/YYYY (French format)
        2. YYYY-MM-DD (ISO format)

        Args:
            date_str: Date string

        Returns:
            date object or None if all formats fail
        """
        if not date_str:
            return None

        if not isinstance(date_str, str):
            date_str = str(date_str)

        date_formats = ["%d/%m/%Y", "%Y-%m-%d"]

        for fmt in date_formats:
            try:
                return datetime.strptime(date_str.strip(), fmt).date()
            except ValueError:
                continue

        return None
