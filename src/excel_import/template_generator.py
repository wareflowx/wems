"""Excel template generator for employee import."""

from pathlib import Path
from typing import Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    raise ImportError("openpyxl is required for Excel template generation. Install it with: pip install openpyxl")

from ui_ctk.constants import (
    CONTRACT_TYPE_CHOICES,
    ROLE_CARISTE,
    ROLE_CHOICES,
    STATUS_ACTIVE,
    STATUS_INACTIVE,
    WORKSPACE_ZONES,
)


class ExcelTemplateGenerator:
    """Generate Excel import template with instructions and validation."""

    # Column definitions
    COLUMNS = [
        "First Name",
        "Last Name",
        "Email",
        "Phone",
        "External ID",
        "Status",
        "Workspace",
        "Role",
        "Contract",
        "Entry Date",
    ]

    # Status dropdown options
    STATUS_OPTIONS = [STATUS_ACTIVE, STATUS_INACTIVE]

    def generate_template(self, output_path: Path) -> None:
        """
        Generate Excel template file with instructions and validation.

        Args:
            output_path: Path where template will be saved
        """
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        # Create instructions sheet
        self._create_instructions_sheet(workbook)

        # Create data sheet
        self._create_data_sheet(workbook)

        # Save workbook
        workbook.save(output_path)
        print(f"[OK] Template generated: {output_path}")

    def _create_instructions_sheet(self, workbook) -> None:
        """Create sheet with import instructions."""
        sheet = workbook.create_sheet("Instructions", 0)

        # Title
        title_cell = sheet["A1"]
        title_cell.value = "Employee Import Template"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # Instructions
        instructions = [
            "",
            "How to use this template:",
            "",
            "1. Fill in the 'Data' sheet with employee information",
            "2. All columns marked with * are required",
            "3. Use dropdown lists for Status, Workspace, Role, and Contract",
            "4. Date format: DD/MM/YYYY (e.g., 15/01/2025)",
            "",
            "Required Columns:",
            "  • First Name* - Employee's first name",
            "  • Last Name* - Employee's last name",
            "  • Status* - Actif or Inactif",
            "  • Workspace* - Employee's work area",
            "  • Role* - Job position",
            "  • Contract* - CDI, CDD, Interim, Alternance",
            "  • Entry Date* - Hire date (DD/MM/YYYY format)",
            "",
            "Optional Columns:",
            "  • Email - Employee email address",
            "  • Phone - Contact phone number",
            "  • External ID - WMS or system reference",
            "",
            "Validation Rules:",
            "  • First/Last name: 1-50 characters, letters only",
            "  • Email: Valid email format if provided",
            "  • Phone: Valid phone format if provided",
            "  • External ID: Must be unique (duplicate will be skipped)",
            "  • Entry Date: Cannot be in the future, must be after 2000",
            "",
            "Import Process:",
            "  1. Download this template",
            "  2. Fill in employee data",
            "  3. Save as .xlsx file",
            "  4. Import via Wareflow EMS application",
            "",
            "Notes:",
            "  • Duplicate External IDs will be skipped (not imported)",
            "  • Invalid entries will show error details",
            "  • Import is done in batches of 100 rows",
            "  • Progress is shown during import",
            "",
            "For support, contact: wareflow@example.com",
        ]

        # Write instructions
        row_num = 3
        for line in instructions:
            sheet[f"A{row_num}"] = line
            row_num += 1

        # Auto-adjust column widths
        sheet.column_dimensions["A"].width = 80

    def _create_data_sheet(self, workbook) -> None:
        """Create data sheet with headers, validation, and example."""
        sheet = workbook.create_sheet("Data", 1)

        # Style definitions
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="FFE6B3", end_color="FFE6B3", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            top=Side(style="thin"), left=Side(style="thin"), bottom=Side(style="thin"), right=Side(style="thin")
        )

        # Create headers
        for col_idx, column_name in enumerate(self.COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            # Add asterisk to required columns
            if self._is_required_column(column_name):
                if column_name != "Entry Date":  # Entry Date already has *
                    cell.value = f"{column_name} *"

            # Auto-fit column width
            sheet.column_dimensions[get_column_letter(col_idx)].width = 18

        # Add data validation for enum columns
        row_num = 1
        for col_idx, column_name in enumerate(self.COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)

            # Status dropdown
            if column_name == "Status":
                dv = DataValidation(type="list", formula1=f'"{",".join(self.STATUS_OPTIONS)}"', allow_blank=False)
                sheet.add_data_validation(dv)
                dv.add(f"{col_letter}{row_num + 1}:{col_letter}{row_num + 1000}")

            # Workspace dropdown
            elif column_name == "Workspace":
                dv = DataValidation(type="list", formula1=f'"{",".join(WORKSPACE_ZONES)}"', allow_blank=False)
                sheet.add_data_validation(dv)
                dv.add(f"{col_letter}{row_num + 1}:{col_letter}{row_num + 1000}")

            # Role dropdown
            elif column_name == "Role":
                dv = DataValidation(type="list", formula1=f'"{",".join(ROLE_CHOICES)}"', allow_blank=False)
                sheet.add_data_validation(dv)
                dv.add(f"{col_letter}{row_num + 1}:{col_letter}{row_num + 1000}")

            # Contract dropdown
            elif column_name == "Contract":
                dv = DataValidation(type="list", formula1=f'"{",".join(CONTRACT_TYPE_CHOICES)}"', allow_blank=False)
                sheet.add_data_validation(dv)
                dv.add(f"{col_letter}{row_num + 1}:{col_letter}{row_num + 1000}")

            # Entry date format
            elif column_name == "Entry Date":
                # Note: Can't easily add date format validation in openpyxl
                # This is handled by our parser
                pass

        # Add example row (row 2) - commented as example
        example_row = {
            "First Name": "Jean",
            "Last Name": "Dupont",
            "Email": "jean.dupont@example.com",
            "Phone": "06 12 34 56 78",
            "External ID": "WMS-001",
            "Status": STATUS_ACTIVE,
            "Workspace": "Zone A",
            "Role": ROLE_CARISTE,
            "Contract": CONTRACT_TYPE_CHOICES[0],
            "Entry Date": "15/01/2025",
        }

        for col_idx, (col_name, value) in enumerate(example_row.items(), start=1):
            cell = sheet.cell(row=2, column=col_idx)
            cell.value = value
            cell.font = Font(italic=True, color="808080")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze header row
        sheet.freeze_panes = "A2"

        # Select the data sheet
        workbook.active = sheet

    @staticmethod
    def _is_required_column(column_name: str) -> bool:
        """Check if column is required."""
        return column_name in ["First Name", "Last Name", "Status", "Workspace", "Role", "Contract", "Entry Date"]

    def generate_sample_file(self, output_path: Path, num_employees: int = 5) -> None:
        """
        Generate a sample Excel file with dummy data for testing.

        Args:
            output_path: Path to save sample file
            num_employees: Number of sample employees to generate
        """
        workbook = Workbook()
        sheet = workbook.active

        # Add headers
        for col_idx, column_name in enumerate(self.COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.font = Font(bold=True)

        # Add sample data
        sample_data = self._generate_sample_data(num_employees)
        row_num = 2

        for employee in sample_data:
            sheet.cell(row_num, 1).value = employee["first_name"]
            sheet.cell(row_num, 2).value = employee["last_name"]
            sheet.cell(row_num, 3).value = employee["email"]
            sheet.cell(row_num, 4).value = employee["phone"]
            sheet.cell(row_num, 5).value = employee["external_id"]
            sheet.cell(row_num, 6).value = employee["status"]
            sheet.cell(row_num, 7).value = employee["workspace"]
            sheet.cell(row_num, 8).value = employee["role"]
            sheet.cell(row_num, 9).value = employee["contract_type"]
            sheet.cell(row_num, 10).value = employee["entry_date"]

            row_num += 1

        # Auto-fit columns
        for col_idx in range(1, 11):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].auto_size = True

        workbook.save(output_path)
        print(f"[OK] Sample file generated: {output_path} with {num_employees} employees")

    def _generate_sample_data(self, count: int) -> List[Dict[str, str]]:
        """Generate sample employee data for testing."""
        first_names = ["Jean", "Marie", "Pierre", "Sophie", "Thomas"]
        last_names = ["Dupont", "Martin", "Bernard", "Richard", "Petit"]

        sample_data = []
        for i in range(count):
            sample_data.append(
                {
                    "first_name": first_names[i % len(first_names)],
                    "last_name": last_names[i % len(last_names)],
                    "email": f"employee{i + 1}@example.com",
                    "phone": f"06 12 34 5{i:02d}",
                    "external_id": f"WMS-{str(i + 1).zfill(3)}",
                    "status": STATUS_ACTIVE if i % 4 != 0 else STATUS_INACTIVE,
                    "workspace": WORKSPACE_ZONES[i % len(WORKSPACE_ZONES)],
                    "role": ROLE_CHOICES[i % len(ROLE_CHOICES)],
                    "contract_type": CONTRACT_TYPE_CHOICES[i % len(CONTRACT_TYPE_CHOICES)],
                    "entry_date": f"{15 + i:02d}/01/2025",
                }
            )

        return sample_data
